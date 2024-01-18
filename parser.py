from bs4 import BeautifulSoup
import requests
import xlsxwriter
import re
import openpyxl

link = 'https://forintek.ru/catalog/printer-etiketok-snbc-btp-7400/'
# print(r.status_code)
# print(r.text)


class Competitors:
    # def __init__(self, counter, article, name, print_method, resolution, print_width, print_speed, memory, interface, old_shtrikh, new_shtrikh, paper_type, paper_roll_width, paper_thicknes, roll_outer_diameter, sleeve_diameter, input_signal, tape_length, tape_sleeve_diameter, operating_mode, storage_condition, size, weight, link):
        # self.article = article
        # self.name = name
        # self.print_method = print_method
        # self.resolution = resolution
        # self.print_width = print_width
        # self.print_speed = print_speed
        # self.memory = memory
        # self.interface = interface
        # self.old_shtrikh = old_shtrikh
        # self.new_shtrikh = new_shtrikh
        # self.paper_type = paper_type
        # self.paper_roll_width = paper_roll_width
        # self.paper_thicknes = paper_thicknes
        # self.roll_outer_diameter = roll_outer_diameter
        # self.sleeve_diameter = sleeve_diameter
        # self.input_signal = input_signal
        # self.tape_length = tape_length
        # self.tape_sleeve_diameter = tape_sleeve_diameter
        # self.operating_mode = operating_mode
        # self.storage_condition = storage_condition
        # self.size = size
        # self.weight = weight
        # self.link = link
        # self.counter = counter

    def __init__(self):
        self.attrs = []

    def steal(self):
        r = requests.get(link)
        soup = BeautifulSoup(r.text, 'lxml')
        name = soup.find_all('h1')
        name = name[0].text


        b = soup.find_all('th')


        article = b[1].text

        self.attrs.append('name')
        self.attrs.append(article)
        self.attrs.append('name')
        self.attrs.append(name)


        c = soup.find_all('td')


        for i in c:
            self.attrs.append(i.text)

        self.attrs.append('dsdd')
        self.attrs.append(link)










    def excel(self):
        workbook = xlsxwriter.Workbook('forintek.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'Артикул')
        worksheet.write(0, 1, 'Наименование')
        worksheet.write(0, 2, 'Метод печати')
        worksheet.write(0, 3, 'Разрешение')
        worksheet.write(0, 4, 'Ширина печати')
        worksheet.write(0, 5, 'Скорость печати')
        worksheet.write(0, 6, 'Память')
        worksheet.write(0, 7, 'Интерфейс')
        worksheet.write(0, 8, '1D штрихкодов')
        worksheet.write(0, 9, '2D штрихкодов')
        worksheet.write(0, 10, 'Тип бумаги')
        worksheet.write(0, 11, 'Ширина рулона бумаги')
        worksheet.write(0, 12, 'Толщина бумаги')
        worksheet.write(0, 13, 'Внешний диаметр рулона')
        worksheet.write(0, 14, 'Диаметр втулки')
        worksheet.write(0, 15, 'Входной сигнал')
        worksheet.write(0, 16, 'Длина ленты')
        worksheet.write(0, 17, 'Диаметр втулки ленты')
        worksheet.write(0, 18, 'Эксплуатационный режим')
        worksheet.write(0, 19, 'Условия хранения')
        worksheet.write(0, 20, 'Размеры')
        worksheet.write(0, 21, 'Вес')
        worksheet.write(0, 22, 'Ссылка')
        worksheet.set_default_row(22)

        bold = workbook.add_format({'bold': True})
        worksheet.set_row(0, 22, bold)
        worksheet.set_column('A:T', 22)
        worksheet.freeze_panes(1, 0)

        workbook.close()

    def safe_date(self):
        wb = openpyxl.load_workbook('forintek.xlsx')
        sheet = wb.active


        x = 1
        for index_i, i in enumerate(self.attrs):

            if index_i % 2 != 0:
                sheet.cell(row=2, column=x).value = i
                x += 1
        wb.save('forintek.xlsx')





x = Competitors()
x.steal()
x.excel()
x.safe_date()





























# url = 'https://meridiant.ru/'
#
# sitemap = 'sitemap.xml'
# sitemap_list = list()
#
# sitemap_url = requests.get(url + sitemap)
# print(sitemap_url)
# sitemap_list.append(sitemap_url)
# print(sitemap_list)
#
# page_html = BeautifulSoup(sitemap.content, 'lxml')
# print(page_html)


# competitors = {
# 'Меридиан':	'https://meridiant.ru/',
# 'Векас': 'https://vekas-automation.ru/',
# 'Тензор': 'https://alfacont.ru/',
# 'Вайландт': 'https://weilandt-elektronik.ru/',
# 'Антарес':	'https://antaresvision.su/',
# 'Трекмарк':	'https://trekmark.ru/',
# 'Оператор ЦРПТ':	'https://crpt.ru/',
# 'Инавтоматика':	'https://inautomatic.ru/',
# 'Альфа технологии': 'http://xn----7sbanjtasdtlb1czat0i.xn--p1ai/',
# 'Контур': 'https://kontur.ru/',
# 'Первый бит': 'https://www.1cbit.ru/',
# 'Эксель': 'https://diar-mark.ru/',
# 'РЭЙ': 'https://chmark.ru/',
# 'ИнСофтРитейл':	'https://insoftretail.ru/',
# 'Мотрум': 'https://www.motrum.ru/',
# 'ОКТО':	'https://okto.ru/',
# 'Ютрейс':	'https://utrace.ru/',
# 'Промышленная маркировка':	'https://www.markprom.ru/',
# # ООО "РБС-Групп"	https://rbsgr.ru/
# # ООО "ИТ-Кластер"	https://it-klaster.com/
# # ООО "Техно Групп"	https://tehnogrupp.com/
# # ООО "ИД Раша"	https://id-russia.ru/
# # ООО "Компания "Тензор" 	https://tensor.ru/
# # АО "Центроинформ"	https://center-inform.ru/
# # ООО "Алкософттрейд"	https://www.alco-dec.ru/
# # Евроконвейер	https://euroconveyor-st.ru/
# # Окзо Ост	https://okzo-ost.ru/
# # КранДеталь	https://kran-dt.ru/
# # ГК "ПОРТ"	https://www.conveyery.ru/
# # Ярославский конвейер	https://gkmash.ru/
# # ООО «ТРАЯНА»	https://zavod-conveyer.ru/
# # OOO "СТиЛ" 	https://konveyery.ru/
# # Satom.ru, ИП Томащик Галина Васильевна	https://moskva.satom.ru/
# # IPC2U	https://ipc2u.ru/
# # ООО «Встраиваемые Системы»	https://empc.ru/
# # Ниеншанц-Автоматика	https://nnz-ipc.ru/
# # ГК "Пром-ПК"	https://prom-pc.ru/
# #  IPC2U	https://aveon.ru/
# # Hualian Machinery Co., Ltd 	https://hmru.ru/
# # Скейл Энтерпрайз, ООО	https://xn--b1afbatd0c9b3b.xn--p1ai/
# # КДМ Трейдинг	https://kdm-trading.ru/
# # Фаспак	https://faspack.ru/
# # Модуль Веса	https://modul-ves.ru/
# # Аспром	https://aspromservis.ru/
# # ООО "Витэк-Автоматика"	https://www.vitec.ru/
# # Оптимус драйв	https://optimusdrive.ru/
# # IPC2U	https://aveon.ru/
# # ООО Видящие машины \Vision Machines 	https://visionmachines.ru/
# # Промфорт	https://promfort.com/
# # Сенсотек	https://sensotek.ru/
# # Hikrobot	https://hikrobot.mallenom.ru/
# # «ЧИП и ДИП» 	https://www.chipdip.ru/
# # РусАвтоматизация	https://rusautomation.ru/
# # Сенсорен Электро, ООО	https://sensoren.ru/
# #  ООО «Эффективное производство» 	https://cnc360.ru/
# # Элрус	https://elrus.ru/
# # Маркджет	https://markjet.ru/
# # ООО "РБС-Групп" (21 в списке)	https://rbs-id.ru/
# # ООО «Упаковочные Системы»	https://packsyst.ru/
# # Профпринт	https://www.profiprint.ru/
# # ООО Арни-Групп	https://print-apply.ru/
# # ИноксДрайв	https://inox-drive.ru/
# # Европактрейд	https://www.pack-euro.ru/
# # Новые Решения	https://oborud-m.ru/
# # Вега Проект 	https://vegatm.ru/
# # Стэн Холдинг	https://stan-upakovka.ru/
# # ПостПромТех	https://postpromteh.ru/
# # Центр КТ	https://shtrih-center.ru/
# # БристольГрупп	https://upakovka-markirovka.ru/
# # Современные технологии	https://www.labelprinter.ru/
# # ООО «Лейблпак»	https://labelpack.ru/
# # Интер Ай Ди	https://interid.ru/
# # Элайтс	smartcode.ru/
# # СТАНДАРТПАК	https://standartpak.ru/
# }










